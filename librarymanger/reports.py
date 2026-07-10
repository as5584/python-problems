from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config import REPORTS_DIR
from models import Book, Issue


def _ensure_reports_dir() -> Path:
    REPORTS_DIR.mkdir(exist_ok=True)
    return REPORTS_DIR


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _book_rows(books: list[Book]) -> list[list]:
    return [
        [
            book.id,
            book.title,
            book.author,
            book.isbn,
            book.total_copies,
            book.available_copies,
            book.total_copies - book.available_copies,
            "Available" if book.available_copies > 0 else "Not Available",
        ]
        for book in books
    ]


def _issue_rows(issues: list[Issue]) -> list[list]:
    return [
        [
            issue.issue_id,
            issue.book_title,
            issue.borrower_name,
            issue.borrower_id,
            issue.issue_date.isoformat(),
            issue.return_date.isoformat() if issue.return_date else "Active",
        ]
        for issue in issues
    ]


def generate_xlsx_report(books: list[Book], issues: list[Issue]) -> Path:
    reports_dir = _ensure_reports_dir()
    output_path = reports_dir / f"library_report_{_timestamp()}.xlsx"

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F5496")
    center = Alignment(horizontal="center", vertical="center")

    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Library Management Report"
    summary["A1"].font = Font(name="Arial", bold=True, size=14)
    summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    total_copies = sum(b.total_copies for b in books)
    available = sum(b.available_copies for b in books)
    active_issues = [i for i in issues if i.return_date is None]

    summary_rows = [
        ("Metric", "Value"),
        ("Total Book Titles", len(books)),
        ("Total Copies", total_copies),
        ("Available Copies", available),
        ("Issued Copies", total_copies - available),
        ("Active Issues", len(active_issues)),
        ("Total Issue Records", len(issues)),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=4):
        summary.cell(row=row_idx, column=1, value=label)
        summary.cell(row=row_idx, column=2, value=value)
        if row_idx == 4:
            for col in (1, 2):
                cell = summary.cell(row=row_idx, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18

    # Books sheet
    books_sheet = wb.create_sheet("Books")
    book_headers = [
        "Book ID",
        "Title",
        "Author",
        "ISBN",
        "Total Copies",
        "Available",
        "Issued",
        "Status",
    ]
    books_sheet.append(book_headers)
    for row in _book_rows(books):
        books_sheet.append(row)

    for cell in books_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for col, width in zip("ABCDEFGH", [10, 35, 25, 18, 14, 12, 10, 16]):
        books_sheet.column_dimensions[col].width = width

    # Issues sheet
    issues_sheet = wb.create_sheet("Issues")
    issue_headers = [
        "Issue ID",
        "Book Title",
        "Borrower Name",
        "Borrower ID",
        "Issue Date",
        "Return Date",
    ]
    issues_sheet.append(issue_headers)
    for row in _issue_rows(issues):
        issues_sheet.append(row)

    for cell in issues_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for col, width in zip("ABCDEF", [10, 35, 25, 15, 14, 14]):
        issues_sheet.column_dimensions[col].width = width

    wb.save(output_path)
    return output_path


def generate_pdf_report(books: list[Book], issues: list[Issue]) -> Path:
    reports_dir = _ensure_reports_dir()
    output_path = reports_dir / f"library_report_{_timestamp()}.pdf"

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=12,
        textColor=colors.HexColor("#2F5496"),
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=16,
    )

    total_copies = sum(b.total_copies for b in books)
    available = sum(b.available_copies for b in books)
    active_issues = [i for i in issues if i.return_date is None]

    elements = [
        Paragraph("Library Management Report", title_style),
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            subtitle_style,
        ),
    ]

    summary_data = [
        ["Total Book Titles", str(len(books))],
        ["Total Copies", str(total_copies)],
        ["Available Copies", str(available)],
        ["Issued Copies", str(total_copies - available)],
        ["Active Issues", str(len(active_issues))],
    ]
    summary_table = Table(summary_data, colWidths=[2.5 * inch, 1.5 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.extend([summary_table, Spacer(1, 0.3 * inch)])

    book_headers = ["ID", "Title", "Author", "ISBN", "Total", "Avail", "Status"]
    book_data = [book_headers] + [
        [
            book.id,
            book.title[:30],
            book.author[:20],
            book.isbn,
            str(book.total_copies),
            str(book.available_copies),
            "Available" if book.available_copies > 0 else "Out",
        ]
        for book in books
    ]
    book_table = Table(
        book_data,
        colWidths=[0.6 * inch, 2.2 * inch, 1.5 * inch, 1.3 * inch, 0.6 * inch, 0.6 * inch, 0.7 * inch],
        repeatRows=1,
    )
    book_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
            ]
        )
    )
    elements.extend(
        [
            Paragraph("Books Inventory", styles["Heading2"]),
            book_table,
            Spacer(1, 0.3 * inch),
        ]
    )

    active = [i for i in issues if i.return_date is None]
    if active:
        issue_headers = ["Issue ID", "Book", "Borrower", "Borrower ID", "Issued"]
        issue_data = [issue_headers] + [
            [
                issue.issue_id,
                issue.book_title[:28],
                issue.borrower_name[:20],
                issue.borrower_id,
                issue.issue_date.isoformat(),
            ]
            for issue in active
        ]
        issue_table = Table(
            issue_data,
            colWidths=[0.8 * inch, 2.5 * inch, 1.5 * inch, 1.2 * inch, 1.0 * inch],
            repeatRows=1,
        )
        issue_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
                ]
            )
        )
        elements.extend(
            [
                Paragraph("Currently Issued Books", styles["Heading2"]),
                issue_table,
            ]
        )
    else:
        elements.append(Paragraph("No books currently issued.", styles["Normal"]))

    doc.build(elements)
    return output_path